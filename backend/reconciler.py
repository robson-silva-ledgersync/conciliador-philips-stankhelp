"""Logica de conciliacao entre planilha Philips e planilha Stankhelp."""

import re
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


@dataclass
class ReconciliationResult:
    philips_count: int = 0
    stankhelp_count: int = 0
    conciliados: list = field(default_factory=list)
    divergencias: list = field(default_factory=list)
    only_philips: list = field(default_factory=list)
    only_stank: list = field(default_factory=list)
    only_philips_swos: set = field(default_factory=set)
    only_stank_swos: set = field(default_factory=set)
    common_swos: set = field(default_factory=set)


# ---------------- Helpers ----------------

def normalize_swo(swo):
    s = str(swo).strip()
    s = re.sub(r'-\d+$', '', s)
    s = re.sub(r'\.0$', '', s)
    return s


def parse_date(val):
    if isinstance(val, datetime):
        return val.strftime('%d.%m.%Y')
    if isinstance(val, str):
        return val.strip()
    return str(val) if val else ''


def normalize_str(val):
    if val is None:
        return ''
    return str(val).strip().upper()


def extract_city(combined):
    if not combined:
        return ''
    s = str(combined).strip().upper()
    return re.split(r'[,/]', s)[0].strip()


def get_field(rec, *substrings):
    for k, v in rec.items():
        if k:
            k_lower = str(k).lower()
            if all(sub in k_lower for sub in substrings):
                return v
    return None


# ---------------- Loaders ----------------

def _extract_month(val) -> str | None:
    """Extrai 'YYYY-MM' de datetime, date ou string DD.MM.YYYY / DD/MM/YYYY."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m')
    if hasattr(val, 'year') and hasattr(val, 'month'):
        return f'{val.year:04d}-{val.month:02d}'
    s = str(val).strip()
    m = re.match(r'(\d{1,2})[./](\d{1,2})[./](\d{4})', s)
    if m:
        day, month, year = m.groups()
        return f'{int(year):04d}-{int(month):02d}'
    return None


def load_philips(
    path: str,
    representante_filter: str | None = None,
    reference_month: str | None = None,
) -> list[dict]:
    """
    Le a planilha base Philips.

    Args:
        path: caminho do arquivo .xlsx
        representante_filter: substring a buscar no campo Representante (ex: 'STANK HELP').
            Se None, retorna todos os representantes.
        reference_month: 'YYYY-MM' para filtrar por Mes Referencia ou Data de Atendimento.
            Se None, retorna todos os meses.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    if 'Reembolso de Despesas' not in wb.sheetnames:
        raise ValueError(
            f"Aba 'Reembolso de Despesas' nao encontrada. "
            f"Abas disponiveis: {wb.sheetnames}"
        )
    ws = wb['Reembolso de Despesas']
    headers = [cell.value for cell in ws[1]]
    rep_upper = representante_filter.strip().upper() if representante_filter else None

    data = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[2] is None:
            continue

        # Filtro por Representante (coluna 1)
        if rep_upper:
            rep = row[1]
            if not rep or rep_upper not in str(rep).upper():
                continue

        # Filtro por mes (coluna 0 = Mes Referencia, fallback coluna 5 = Data de Atendimento)
        if reference_month:
            month = _extract_month(row[0]) or _extract_month(row[5])
            if month != reference_month:
                continue

        data.append(dict(zip(headers, row)))
    return data


def load_stankhelp(
    path: str,
    reference_month: str | None = None,
) -> list[dict]:
    """
    Le a planilha do representante (Stankhelp).

    Args:
        path: caminho do arquivo .xlsx
        reference_month: 'YYYY-MM' para filtrar por Data de Atendimento.
            Se None, retorna todos os registros com SWO (incluindo historicos).

    A planilha pode ter linhas vazias intercaladas (dropdowns de validacao) e
    registros historicos de meses anteriores no final. O loader le TODAS as
    linhas com SWO valido e filtra por mes se solicitado.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    if 'Reembolso' not in wb.sheetnames:
        raise ValueError(
            f"Aba 'Reembolso' nao encontrada. "
            f"Abas disponiveis: {wb.sheetnames}"
        )
    ws = wb['Reembolso']
    headers_raw = [cell.value for cell in ws[10]]
    headers = [
        (h.strip().replace('\n', ' ') if h else h) for h in headers_raw
    ]
    data = []
    for row in ws.iter_rows(min_row=11, max_row=ws.max_row, values_only=True):
        swo_val = row[4]
        if swo_val is None or not str(swo_val).strip():
            continue

        # Filtro por mes (coluna 6 = Data de Atendimento)
        if reference_month:
            month = _extract_month(row[6])
            if month != reference_month:
                continue

        data.append(dict(zip(headers, row)))
    return data


# ---------------- Core reconciliation ----------------

TIPO_MAP = {
    'GARANTIA PADRÃO': 'GARANTIA',
    'GARANTIA PADRAO': 'GARANTIA',
    'CONTRATO DE MANUTENÇÃO': 'CONTRATO',
    'CONTRATO DE MANUTENCAO': 'CONTRATO',
    'GARANTIA ESTENDIDA': 'GARANTIA EST',
    'INSTALAÇÃO': 'INSTALACAO',
    'INSTALACAO': 'INSTALACAO',
    'SERVIÇO AVULSO': 'SERVICO AVULSO',
    'SERVICO AVULSO': 'SERVICO AVULSO',
    'SERVIÇO AVULSO (NÃO FATURADO)': 'SERVICO AVULSO',
    'LIBERADO EM GARANTIA': 'GARANTIA',
}


def _build_matched_record(swo, p, s, diffs, diff_fields, p_date, s_date, p_serial):
    return {
        'SWO': swo,
        'SWO Stankhelp': s.get('SWO', ''),
        'Cliente Philips': p.get('Customer Name', ''),
        'Cliente Stankhelp': get_field(s, 'cliente') or s.get('   Cliente', '') or '',
        'Serial': p_serial,
        'Equipamento Philips': p.get('Equipamento', ''),
        'Equipamento Stankhelp': get_field(s, 'equipamento') or '',
        'Tipo Atend Philips': p.get('Tipo de Atendimento', ''),
        'Tipo Atend Stankhelp': s.get('Tipo de Atendimento', ''),
        'Data Philips': p_date,
        'Data Stankhelp': s_date,
        'Distância KM': p.get('Distância KM', ''),
        'Outras Despesas': p.get('Outras Despesas', ''),
        'Quilometragem': p.get('Quilometragem', ''),
        'Hospedagem': p.get('Hospedagem', ''),
        'Reembolso Total': p.get('Reembolso total', ''),
        'MDO (R$)': p.get('MDO (R$)', ''),
        'Técnico': get_field(s, 'cnico') or '',
        'Atividade Philips': p.get('Atividade', ''),
        'Atividade Stankhelp': get_field(s, 'corretiva') or '',
        'Cidade Philips': p.get('Cidade/Estado Destino', ''),
        'Cidade Stankhelp': get_field(s, 'cidade', 'destino') or '',
        'Contrato': p.get('Contrato', ''),
        'Sales Document': p.get('Sales Document', ''),
        'Observações': get_field(p, 'observa') or '',
        '_diff_fields': diff_fields,
        '_diffs': diffs,
    }


def reconcile(philips_data: list[dict], stank_data: list[dict]) -> ReconciliationResult:
    result = ReconciliationResult(
        philips_count=len(philips_data),
        stankhelp_count=len(stank_data),
    )

    philips_by_swo = defaultdict(list)
    for rec in philips_data:
        philips_by_swo[normalize_swo(rec.get('SWO', ''))].append(rec)

    stank_by_swo = defaultdict(list)
    for rec in stank_data:
        stank_by_swo[normalize_swo(rec.get('SWO', ''))].append(rec)

    philips_swos = set(philips_by_swo.keys())
    stank_swos = set(stank_by_swo.keys())

    result.only_philips_swos = philips_swos - stank_swos
    result.only_stank_swos = stank_swos - philips_swos
    result.common_swos = philips_swos & stank_swos

    # ---- Match records ----
    for swo in sorted(result.common_swos):
        p_recs = philips_by_swo[swo]
        s_recs = stank_by_swo[swo]
        used_s_indices = set()  # evita que 2 p_recs matchem o mesmo s_rec

        for p in p_recs:
            matched = False
            p_serial = normalize_str(get_field(p, 'rie') or '')

            for s_idx, s in enumerate(s_recs):
                if s_idx in used_s_indices:
                    continue
                s_serial = normalize_str(get_field(s, 'rie') or '')
                if p_serial != s_serial:
                    continue

                used_s_indices.add(s_idx)
                matched = True
                diffs = []
                diff_fields = set()

                p_date = parse_date(p.get('Data de Atendimento', ''))
                s_date = str(s.get('Data de Atendimento', '') or '').strip()

                # Cidade
                p_city = extract_city(p.get('Cidade/Estado Destino', ''))
                s_city = normalize_str(get_field(s, 'cidade', 'destino') or '')
                if p_city and s_city and p_city != s_city:
                    if p_city not in s_city and s_city not in p_city:
                        diffs.append(f'Cidade: Philips={p_city} vs Stank={s_city}')
                        diff_fields.add('cidade')

                # Tipo
                p_tipo = normalize_str(p.get('Tipo de Atendimento', ''))
                s_tipo = normalize_str(s.get('Tipo de Atendimento', ''))
                if TIPO_MAP.get(p_tipo, p_tipo) != TIPO_MAP.get(s_tipo, s_tipo):
                    diffs.append(f'Tipo: Philips={p_tipo} vs Stank={s_tipo}')
                    diff_fields.add('tipo')

                # Atividade
                p_ativ = normalize_str(p.get('Atividade', ''))
                s_ativ = normalize_str(get_field(s, 'corretiva') or '')
                if p_ativ and s_ativ:
                    if p_ativ not in s_ativ and s_ativ not in p_ativ:
                        if 'INSTAL' not in p_ativ and 'INSTAL' not in s_ativ:
                            diffs.append(f'Atividade: Philips={p_ativ} vs Stank={s_ativ}')
                            diff_fields.add('atividade')

                row_data = _build_matched_record(
                    swo, p, s, diffs, diff_fields, p_date, s_date, p_serial
                )
                if diffs:
                    result.divergencias.append(row_data)
                else:
                    result.conciliados.append(row_data)
                break

            if not matched:
                # Pega o primeiro s_rec ainda nao consumido. Se todos ja foram
                # consumidos, usa s_recs[0] apenas como placeholder visual.
                fallback_idx = next(
                    (i for i in range(len(s_recs)) if i not in used_s_indices),
                    None,
                )
                if fallback_idx is not None:
                    s = s_recs[fallback_idx]
                    used_s_indices.add(fallback_idx)
                else:
                    s = s_recs[0]
                row_data = _build_matched_record(
                    swo, p, s,
                    ['Numero de serie nao encontrado no Stankhelp'],
                    {'serial'},
                    parse_date(p.get('Data de Atendimento', '')),
                    str(s.get('Data de Atendimento', '') or ''),
                    normalize_str(get_field(p, 'rie') or ''),
                )
                result.divergencias.append(row_data)

        # Stankhelp recs desse SWO que nao foram consumidos
        for s_idx, s in enumerate(s_recs):
            if s_idx in used_s_indices:
                continue
            result.only_stank.append({
                'SWO': s.get('SWO', ''),
                'Cliente': get_field(s, 'cliente') or s.get('   Cliente', '') or '',
                'Data Atendimento': str(s.get('Data de Atendimento', '') or ''),
                'Estado Destino': get_field(s, 'estado', 'destino') or '',
                'Cidade Destino': get_field(s, 'cidade', 'destino') or '',
                'Tipo Atendimento': s.get('Tipo de Atendimento', ''),
                'Atividade': get_field(s, 'corretiva') or '',
                'Equipamento': get_field(s, 'equipamento') or '',
                'Número de Série': get_field(s, 'rie') or '',
                'Técnico': get_field(s, 'cnico') or '',
            })

    # ---- Only Philips ----
    for swo in sorted(result.only_philips_swos):
        for p in philips_by_swo[swo]:
            result.only_philips.append({
                'SWO': swo,
                'Cliente': p.get('Customer Name', ''),
                'Data Atendimento': parse_date(p.get('Data de Atendimento', '')),
                'Cidade Destino': p.get('Cidade/Estado Destino', ''),
                'Atividade': p.get('Atividade', ''),
                'Tipo Atendimento': p.get('Tipo de Atendimento', ''),
                'Equipamento': p.get('Equipamento', ''),
                'Número de Série': get_field(p, 'rie') or '',
                'Distância KM': p.get('Distância KM', ''),
                'Outras Despesas': p.get('Outras Despesas', ''),
                'Quilometragem': p.get('Quilometragem', ''),
                'Hospedagem': p.get('Hospedagem', ''),
                'Reembolso Total': p.get('Reembolso total', ''),
                'MDO (R$)': p.get('MDO (R$)', ''),
                'Observações': get_field(p, 'observa') or '',
            })

    # ---- Only Stankhelp ----
    for swo in sorted(result.only_stank_swos):
        for s in stank_by_swo[swo]:
            result.only_stank.append({
                'SWO': s.get('SWO', ''),
                'Cliente': get_field(s, 'cliente') or s.get('   Cliente', '') or '',
                'Data Atendimento': str(s.get('Data de Atendimento', '') or ''),
                'Estado Destino': get_field(s, 'estado', 'destino') or '',
                'Cidade Destino': get_field(s, 'cidade', 'destino') or '',
                'Tipo Atendimento': s.get('Tipo de Atendimento', ''),
                'Atividade': get_field(s, 'corretiva') or '',
                'Equipamento': get_field(s, 'equipamento') or '',
                'Número de Série': get_field(s, 'rie') or '',
                'Técnico': get_field(s, 'cnico') or '',
            })

    return result


# ---------------- Excel writer ----------------

HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
PHILIPS_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
STANK_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
DIFF_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
DIFF_FONT = Font(bold=True, color='9C0006')
VALOR_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
SECTION_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)


def _write_simple_sheet(ws, title, headers, data_rows):
    ws.title = title
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER
    for row_idx, row_data in enumerate(data_rows, 2):
        for col_idx, h in enumerate(headers, 1):
            val = row_data.get(h, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True)
    for col_idx, h in enumerate(headers, 1):
        max_len = len(str(h))
        for row_data in data_rows:
            val = str(row_data.get(h, ''))
            max_len = max(max_len, min(len(val), 40))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_len + 3


def write_report(result: ReconciliationResult, output_path: str):
    wb = openpyxl.Workbook()

    # Sheet 1: Resumo
    ws_resumo = wb.active
    ws_resumo.title = 'Resumo'
    summary = [
        ['Métrica', 'Quantidade'],
        ['Total registros Philips', result.philips_count],
        ['Total registros Stankhelp', result.stankhelp_count],
        ['SWOs em comum', len(result.common_swos)],
        ['SWOs somente Philips', len(result.only_philips_swos)],
        ['SWOs somente Stankhelp', len(result.only_stank_swos)],
        ['Registros conciliados (sem divergência)', len(result.conciliados)],
        ['Registros com divergências', len(result.divergencias)],
        ['', ''],
        ['SWOs somente na Philips', ', '.join(sorted(result.only_philips_swos))],
        ['SWOs somente no Stankhelp', ', '.join(sorted(result.only_stank_swos))],
    ]
    for row_idx, row_vals in enumerate(summary, 1):
        for col_idx, val in enumerate(row_vals, 1):
            cell = ws_resumo.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            if row_idx == 1:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
            cell.alignment = Alignment(wrap_text=True)
    ws_resumo.column_dimensions['A'].width = 45
    ws_resumo.column_dimensions['B'].width = 60

    # Sheet 2: Somente Philips
    ws_sp = wb.create_sheet()
    sp_headers = [
        'SWO', 'Cliente', 'Data Atendimento', 'Cidade Destino', 'Atividade',
        'Tipo Atendimento', 'Equipamento', 'Número de Série',
        'Distância KM', 'Outras Despesas', 'Quilometragem', 'Hospedagem',
        'Reembolso Total', 'MDO (R$)', 'Observações',
    ]
    _write_simple_sheet(ws_sp, 'Somente Philips', sp_headers, result.only_philips)

    # Sheet 3: Somente Stankhelp
    ws_ss = wb.create_sheet()
    ss_headers = [
        'SWO', 'Cliente', 'Data Atendimento', 'Estado Destino', 'Cidade Destino',
        'Tipo Atendimento', 'Atividade', 'Equipamento', 'Número de Série', 'Técnico',
    ]
    _write_simple_sheet(ws_ss, 'Somente Stankhelp', ss_headers, result.only_stank)

    # Sheet 4: Divergencias (lado a lado)
    ws_div = wb.create_sheet()
    ws_div.title = 'Divergências'

    div_cols = [
        ('Fonte', 12), ('SWO', 16), ('Cliente', 32), ('Nº Série', 16),
        ('Equipamento', 28), ('Tipo Atendimento', 20), ('Atividade', 20),
        ('Cidade Destino', 20), ('Data Atendimento', 15),
        ('Distância KM', 13), ('Outras Desp.', 13), ('Quilometragem', 14),
        ('Hospedagem', 12), ('Reembolso Total', 15), ('MDO (R$)', 12),
        ('Contrato', 18), ('Técnico', 14), ('Observações', 30),
    ]

    for col_idx, (title, width) in enumerate(div_cols, 1):
        cell = ws_div.cell(row=1, column=col_idx, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
        ws_div.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    diff_col_map = {'tipo': 6, 'atividade': 7, 'cidade': 8, 'serial': 4}
    r = 2
    for rec in result.divergencias:
        diff_fields = rec.get('_diff_fields', set())

        p_vals = [
            'PHILIPS', rec.get('SWO', ''), rec.get('Cliente Philips', ''),
            rec.get('Serial', ''), rec.get('Equipamento Philips', ''),
            rec.get('Tipo Atend Philips', ''), rec.get('Atividade Philips', ''),
            rec.get('Cidade Philips', ''), rec.get('Data Philips', ''),
            rec.get('Distância KM', ''), rec.get('Outras Despesas', ''),
            rec.get('Quilometragem', ''), rec.get('Hospedagem', ''),
            rec.get('Reembolso Total', ''), rec.get('MDO (R$)', ''),
            rec.get('Contrato', ''), '', rec.get('Observações', ''),
        ]
        for col_idx, val in enumerate(p_vals, 1):
            cell = ws_div.cell(row=r, column=col_idx, value=val)
            cell.fill = PHILIPS_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True)
            if col_idx in (10, 11, 12, 13, 14, 15) and val not in (None, '', 0):
                cell.fill = VALOR_FILL
                cell.number_format = '#,##0.00'
        ws_div.cell(row=r, column=1).font = Font(bold=True, color='1F4E79')
        for fk, col in diff_col_map.items():
            if fk in diff_fields:
                ws_div.cell(row=r, column=col).fill = DIFF_FILL
                ws_div.cell(row=r, column=col).font = DIFF_FONT
        r += 1

        s_vals = [
            'STANKHELP', rec.get('SWO Stankhelp', ''), rec.get('Cliente Stankhelp', ''),
            rec.get('Serial', ''), rec.get('Equipamento Stankhelp', ''),
            rec.get('Tipo Atend Stankhelp', ''), rec.get('Atividade Stankhelp', ''),
            rec.get('Cidade Stankhelp', ''), rec.get('Data Stankhelp', ''),
            '', '', '', '', '', '', '', rec.get('Técnico', ''), '',
        ]
        for col_idx, val in enumerate(s_vals, 1):
            cell = ws_div.cell(row=r, column=col_idx, value=val)
            cell.fill = STANK_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True)
        ws_div.cell(row=r, column=1).font = Font(bold=True, color='375623')
        for fk, col in diff_col_map.items():
            if fk in diff_fields:
                ws_div.cell(row=r, column=col).fill = DIFF_FILL
                ws_div.cell(row=r, column=col).font = DIFF_FONT
        r += 1

        for col_idx in range(1, len(div_cols) + 1):
            cell = ws_div.cell(row=r, column=col_idx, value='')
            cell.fill = SECTION_FILL
            cell.border = Border(bottom=Side(style='thin', color='BFBFBF'))
        r += 1

    # Sheet 5: Conciliados
    ws_conc = wb.create_sheet()
    conc_headers = [
        'SWO', 'Cliente', 'Nº Série', 'Equipamento', 'Tipo Atendimento',
        'Data Stankhelp', 'Distância KM', 'Outras Desp.', 'Quilometragem',
        'Hospedagem', 'Reembolso Total', 'MDO (R$)', 'Contrato', 'Técnico', 'Observações',
    ]
    conc_simple = []
    for rec in result.conciliados:
        conc_simple.append({
            'SWO': rec.get('SWO', ''),
            'Cliente': rec.get('Cliente Philips', ''),
            'Nº Série': rec.get('Serial', ''),
            'Equipamento': rec.get('Equipamento Philips', ''),
            'Tipo Atendimento': rec.get('Tipo Atend Philips', ''),
            'Data Stankhelp': rec.get('Data Stankhelp', ''),
            'Distância KM': rec.get('Distância KM', ''),
            'Outras Desp.': rec.get('Outras Despesas', ''),
            'Quilometragem': rec.get('Quilometragem', ''),
            'Hospedagem': rec.get('Hospedagem', ''),
            'Reembolso Total': rec.get('Reembolso Total', ''),
            'MDO (R$)': rec.get('MDO (R$)', ''),
            'Contrato': rec.get('Contrato', ''),
            'Técnico': rec.get('Técnico', ''),
            'Observações': rec.get('Observações', ''),
        })
    _write_simple_sheet(ws_conc, 'Conciliados', conc_headers, conc_simple)

    wb.save(output_path)
